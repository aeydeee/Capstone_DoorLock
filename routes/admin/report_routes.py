from openpyxl import Workbook
from flask import Response
import io
import os
from datetime import datetime

from flask import Blueprint, render_template, request, send_file, Response, flash, redirect, url_for
from flask_login import login_required

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape, GOV_LEGAL
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import TableStyle, Table, Paragraph, SimpleDocTemplate, Spacer

from decorators import cspc_acc_required, admin_required
from models import User, ReportLog

report_bp = Blueprint('report', __name__)


def export_faculty_excel(reports):
    # Create an Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Faculty Reports"

    # Write the Excel header
    ws.append(['DATE', 'NAME', 'SCHOOL ID', 'TIME', 'STATUS'])

    # Write the report data for each faculty report
    for report in reports:
        ws.append([
            report.timestamp.strftime('%b. %d, %Y') if report.timestamp else '',
            report.name.title() if report.role.lower() == 'faculty' else '',
            report.school_id.upper() if report.role.lower() == 'faculty' else '',
            report.timestamp.strftime('%I:%M %p') if report.timestamp else '',
            report.status.title()
        ])

    # Save the workbook to a binary buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Return the Excel file to the client
    return Response(buffer, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment;filename=faculty_report.xlsx"})


def export_admin_excel(reports):
    # Create an Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Admin Reports"

    # Write the Excel header
    ws.append(['DATE', 'NAME', 'SCHOOL ID', 'TIME', 'STATUS'])

    # Write the report data for each admin report
    for report in reports:
        ws.append([
            report.timestamp.strftime('%b. %d, %Y') if report.timestamp else '',
            report.name.title() if report.role.lower() == 'admin' else '',
            report.school_id.upper() if report.role.lower() == 'admin' else '',
            report.timestamp.strftime('%I:%M %p') if report.timestamp else '',
            report.status.title()
        ])

    # Save the workbook to a binary buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Return the Excel file to the client
    return Response(buffer, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment;filename=admin_report.xlsx"})


def export_faculty_admin_excel(reports):
    # Create an Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Faculty/Admin Reports"

    # Write the Excel header
    ws.append(['DATE', 'NAME', 'SCHOOL ID', 'TIME', 'STATUS'])

    # Write the report data for each faculty/admin report
    for report in reports:
        name = report.name.title() if report.role.lower() == 'faculty' else report.name.title()
        school_id = report.school_id.upper() if report.role.lower() == 'faculty' else report.school_id.upper()

        ws.append([
            report.timestamp.strftime('%b. %d, %Y') if report.timestamp else '',
            name,
            school_id,
            report.timestamp.strftime('%I:%M %p') if report.timestamp else '',
            report.status.title()
        ])

    # Save the workbook to a binary buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Return the Excel file to the client
    return Response(buffer, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment;filename=faculty_admin_report.xlsx"})


def add_header_footer(canvas, doc, is_first_page, start_date=None, end_date=None):
    canvas.saveState()

    # Header Section - different logic for first page
    if is_first_page:
        # Left side logo
        logo_path = os.path.join(os.getcwd(), 'static', 'images', 'logo', 'cspc.png')
        if os.path.exists(logo_path):
            canvas.drawImage(logo_path, 65, doc.height + 40, width=50, height=50)  # Adjusted logo size and position

        # Left side text
        canvas.setFont("Helvetica", 10)
        canvas.drawString(120, doc.height + 75, "Republic of the Philippines")  # Adjusted header text position
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawString(120, doc.height + 60, "CAMARINES SUR POLYTECHNIC COLLEGES")
        canvas.setFont("Helvetica", 10)
        canvas.drawString(120, doc.height + 45, "Nabua, Camarines Sur")
        canvas.setFont("Helvetica", 6)
        canvas.drawString(70, doc.height + 30, "ISO 9001:2015")

        # Right side logo
        new_logo_path = os.path.join(os.getcwd(), 'static', 'images', 'logo', 'ccs-logo.png')
        if os.path.exists(new_logo_path):
            canvas.drawImage(new_logo_path, doc.width - 1, doc.height + 40, width=50,
                             height=50)  # Align to match left logo size

        # Right side text (College of Computer Studies)
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawString(doc.width - 215, doc.height + 60,
                          "COLLEGE OF COMPUTER STUDIES")  # Text aligned to the left of the logo

        # Horizontal line after the header
        canvas.setLineWidth(1)
        canvas.setStrokeColor(colors.black)
        canvas.line(60, doc.height + 25, doc.width + 45, doc.height + 25)  # Draw a horizontal line after the header

    else:
        # Adjusted header for the 2nd and subsequent pages
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawString(65, doc.height + 60,
                          "CAMARINES SUR POLYTECHNIC COLLEGES, DOOR ACCESS REPORT LOGS - CONTINUED")

    # Add the italic footer text
    canvas.setFont("Helvetica-Oblique", 9)  # Setting the font to italic
    # canvas.drawString(60, 50, "System-generated report. No Signature required.")  # Adjust the y-position as needed

    # Footer Section with horizontal line
    canvas.setLineWidth(1)
    canvas.setStrokeColor(colors.black)
    canvas.line(60, 40, doc.width + 45, 40)  # Horizontal line for the footer
    canvas.setFont("Helvetica", 10)

    # Format the effectivity date based on the provided start and end dates
    if start_date and end_date:
        effectivity_date = f"From {start_date.strftime('%Y-%m-%d')} - To {end_date.strftime('%Y-%m-%d')}"
    elif start_date:
        effectivity_date = f"From {start_date.strftime('%Y-%m-%d')}"
    elif end_date:
        effectivity_date = f"To {end_date.strftime('%Y-%m-%d')}"
    else:
        effectivity_date = "All records across all dates."

    canvas.drawString(60, 30, effectivity_date)

    # canvas.drawCentredString(doc.width // 1.8, 30, "Rev. 0")
    canvas.drawRightString(doc.width + 45, 30, f"Page {doc.page}")

    canvas.restoreState()


def export_faculty_pdf(reports, start_date=None, end_date=None):
    # Create a PDF buffer
    buffer = io.BytesIO()

    # Create a document template with the desired page size and margins
    doc = SimpleDocTemplate(buffer, pagesize=landscape(GOV_LEGAL),
                            leftMargin=0.75 * inch, rightMargin=0.75 * inch,
                            topMargin=1 * inch,
                            bottomMargin=0.75 * inch)

    # Prepare the stylesheet and elements container
    styles = getSampleStyleSheet()
    elements = []

    # Add a large spacer before the title to push it down
    elements.append(Spacer(1, 30))  # Spacer to push the title down

    # Add the document title at the top
    title = Paragraph("<b>DOOR ACCESS FACULTY REPORT LOGS</b>", styles['Title'])
    title.hAlign = 'CENTER'
    elements.append(title)

    # Add another spacer after the title
    elements.append(Spacer(1, 10))

    # Prepare data for the table (header row + report rows)
    table_header = ['DATE', 'NAME', 'SCHOOL ID', 'TIME', 'STATUS']

    data = [[Paragraph(cell, styles['Normal']) for cell in table_header]]  # Wrap header cells in Paragraph

    for report in reports:
        row = [
            Paragraph(report.timestamp.strftime('%b. %d, %Y') if report.timestamp else '', styles['Normal']),
            Paragraph(report.name.title(), styles['Normal']),
            Paragraph(report.school_id.upper(), styles['Normal']),
            Paragraph(report.timestamp.strftime('%I:%M %p') if report.timestamp else '', styles['Normal']),
            Paragraph(report.status.title(), styles['Normal']),
        ]
        data.append(row)

    # Define the table for the PDF
    table = Table(data, colWidths=[70, 150, 100, 70, 200], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.white),  # Header background
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # Header text color
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # Data background
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),  # Data text color
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Grid color
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
    ]))

    # Add the table to the elements list
    elements.append(table)

    # Define page header and footer
    def on_first_page(canvas, doc):
        add_header_footer(canvas, doc, is_first_page=True, start_date=start_date, end_date=end_date)

    def on_later_pages(canvas, doc):
        add_header_footer(canvas, doc, is_first_page=False, start_date=start_date, end_date=end_date)

    # Build the document with conditional page elements
    doc.build(elements, onFirstPage=on_first_page, onLaterPages=on_later_pages)

    # Return the PDF to the client
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='report_records.pdf', mimetype='application/pdf')


def export_admin_pdf(reports, start_date=None, end_date=None):
    # Create a PDF buffer
    buffer = io.BytesIO()

    # Create a document template with the desired page size and margins
    doc = SimpleDocTemplate(buffer, pagesize=landscape(GOV_LEGAL),
                            leftMargin=0.75 * inch, rightMargin=0.75 * inch,
                            topMargin=1 * inch,
                            bottomMargin=0.75 * inch)

    # Prepare the stylesheet and elements container
    styles = getSampleStyleSheet()
    elements = []

    # Add a large spacer before the title to push it down
    elements.append(Spacer(1, 30))  # Spacer to push the title down

    # Add the document title at the top
    title = Paragraph("<b>DOOR ACCESS FACULTY & ADMIN REPORT LOGS</b>", styles['Title'])
    title.hAlign = 'CENTER'
    elements.append(title)

    # Add another spacer after the title
    elements.append(Spacer(1, 10))

    # Prepare data for the table (header row + report rows)
    table_header = ['DATE', 'NAME', 'SCHOOL ID', 'TIME', 'STATUS']

    data = [[Paragraph(cell, styles['Normal']) for cell in table_header]]  # Wrap header cells in Paragraph

    for report in reports:
        row = [
            Paragraph(report.timestamp.strftime('%b. %d, %Y') if report.timestamp else '', styles['Normal']),
            Paragraph(report.name.title(), styles['Normal']),
            Paragraph(report.school_id.upper(), styles['Normal']),
            Paragraph(report.timestamp.strftime('%I:%M %p') if report.timestamp else '', styles['Normal']),
            Paragraph(report.status.title(), styles['Normal']),
        ]
        data.append(row)

    # Define the table for the PDF
    table = Table(data, colWidths=[70, 150, 100, 70, 200], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.white),  # Header background
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # Header text color
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # Data background
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),  # Data text color
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Grid color
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
    ]))

    # Add the table to the elements list
    elements.append(table)

    # Define page header and footer
    def on_first_page(canvas, doc):
        add_header_footer(canvas, doc, is_first_page=True, start_date=start_date, end_date=end_date)

    def on_later_pages(canvas, doc):
        add_header_footer(canvas, doc, is_first_page=False, start_date=start_date, end_date=end_date)

    # Build the document with conditional page elements
    doc.build(elements, onFirstPage=on_first_page, onLaterPages=on_later_pages)

    # Return the PDF to the client
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='report_records.pdf', mimetype='application/pdf')


def export_faculty_admin_pdf(reports, start_date=None, end_date=None):
    # Create a PDF buffer
    buffer = io.BytesIO()

    # Create a document template with the desired page size and margins
    doc = SimpleDocTemplate(buffer, pagesize=landscape(GOV_LEGAL),
                            leftMargin=0.75 * inch, rightMargin=0.75 * inch,
                            topMargin=1 * inch,
                            bottomMargin=0.75 * inch)

    # Prepare the stylesheet and elements container
    styles = getSampleStyleSheet()
    elements = []

    # Add a large spacer before the title to push it down
    elements.append(Spacer(1, 30))  # Spacer to push the title down

    # Add the document title at the top
    title = Paragraph("<b>FACULTIES DOOR ACCESS REPORT LOGS</b>", styles['Title'])
    title.hAlign = 'CENTER'
    elements.append(title)

    # Add another spacer after the title
    elements.append(Spacer(1, 10))

    # Prepare data for the table (header row + report rows)
    table_header = ['DATE', 'NAME', 'SCHOOL ID', 'TIME', 'STATUS']

    data = [[Paragraph(cell, styles['Normal']) for cell in table_header]]  # Wrap header cells in Paragraph

    for report in reports:
        row = [
            Paragraph(report.timestamp.strftime('%b. %d, %Y') if report.timestamp else '', styles['Normal']),
            Paragraph(
                report.name.title() if report.role.lower() == 'faculty' else report.name.title(),
                styles['Normal']),
            Paragraph(
                report.school_id.upper() if report.role.lower() == 'faculty' else report.school_id.upper(),
                styles['Normal']),
            Paragraph(report.timestamp.strftime('%I:%M %p') if report.timestamp else '', styles['Normal']),
            Paragraph(report.status.title(), styles['Normal']),
        ]

        data.append(row)

    # Define the table for the PDF
    table = Table(data, colWidths=[70, 150, 100, 70, 200], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.white),  # Header background
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # Header text color
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # Data background
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),  # Data text color
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Grid color
        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
    ]))

    # Add the table to the elements list
    elements.append(table)

    # Define page header and footer
    def on_first_page(canvas, doc):
        add_header_footer(canvas, doc, is_first_page=True, start_date=start_date, end_date=end_date)

    def on_later_pages(canvas, doc):
        add_header_footer(canvas, doc, is_first_page=False, start_date=start_date, end_date=end_date)

    # Build the document with conditional page elements
    doc.build(elements, onFirstPage=on_first_page, onLaterPages=on_later_pages)

    # Return the PDF to the client
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='report_records.pdf', mimetype='application/pdf')


@report_bp.route('/report-generation', methods=['GET', 'POST'])
@login_required
@cspc_acc_required
@admin_required
def report_generation():
    # Check for export requests
    export_type = request.args.get('export')

    # Handle date filters from query parameters
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    role_filter = request.args.get('role')  # Get role filter from query params

    # Parse the start and end date if provided
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None

    reports_query = ReportLog.query

    # Apply date filters if present
    if start_date:
        reports_query = reports_query.filter(ReportLog.timestamp >= start_date)
    if end_date:
        # Set time to end of the day for the end date filter
        end_date = end_date.replace(hour=23, minute=59, second=59)
        reports_query = reports_query.filter(ReportLog.timestamp <= end_date)

    # Apply role filter if provided
    if role_filter:
        if role_filter == 'FacultyAdmin':
            reports_query = reports_query.join(User, ReportLog.user_id == User.id).filter(
                User.role.in_(['Faculty', 'Admin']))
        else:
            reports_query = reports_query.join(ReportLog.user).filter_by(role=role_filter)

    # Fetch the reports after applying the filters
    reports = reports_query.all()

    # Handle export requests with the filtered data
    if export_type == 'excel' or export_type == 'pdf':
        print(f"Filtered ReportLogs for Export: {reports}")  # Debugging export
        if export_type == 'excel':
            if role_filter == 'Faculty':
                return export_faculty_excel(reports)
            elif role_filter == 'Admin':
                return export_admin_excel(reports)
            elif role_filter == 'FacultyAdmin':
                return export_faculty_admin_excel(reports)
            else:
                flash(
                    'Export Generation is only available for Faculty and Admin. Please proceed to the Attendance Report Logs Page.')
                return redirect(url_for('report.report_generation'))
        elif export_type == 'pdf':
            if role_filter == 'Faculty':
                return export_faculty_pdf(reports)
            elif role_filter == 'Admin':
                return export_admin_pdf(reports)
            elif role_filter == 'FacultyAdmin':
                return export_faculty_admin_pdf(reports)
            else:
                flash(
                    'Export Generation is only available for Faculty and Admin. Please proceed to the Attendance Report Logs Page.')
                return redirect(url_for('report.report_generation'))

    return render_template('report/report.html', reports=reports)
