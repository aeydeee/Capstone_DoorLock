import os
import re
import uuid
from datetime import datetime, timedelta

import io
import os
import re
import pandas as pd
import pytz
from fuzzywuzzy import fuzz
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import letter, landscape, GOV_LEGAL, legal
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import TableStyle, Table, Paragraph, Image, SimpleDocTemplate, Spacer, BaseDocTemplate, Frame, \
    PageTemplate, KeepInFrame

from flask import Blueprint, request, render_template, flash, redirect, url_for, jsonify, abort, session, current_app, \
    make_response, send_file
from flask_login import login_required, current_user, logout_user
from sqlalchemy.exc import SQLAlchemyError, IntegrityError
from sqlalchemy.orm import aliased
# from werkzeug.security import generate_password_hash
from werkzeug.utils import secure_filename

from app import db
from decorators import cspc_acc_required, faculty_required, own_faculty_account_required, check_totp_verified
from models import Faculty, User, Student, Attendance, Schedule, Course, \
    student_course_association, faculty_course_association, Section, FacultyCourseSchedule, Program, YearLevel, \
    Semester, SchoolYear
from webforms.attendance_form import SelectScheduleForm
from webforms.delete_form import DeleteForm
from webforms.faculty_acc_form import AttendanceStatusForm
from webforms.faculty_form import FacultyForm

faculty_acc_bp = Blueprint('faculty_acc', __name__)


# EXPORT ROUTES
def export_excel_detail(attendances):
    # Convert attendance records to a list of dictionaries
    attendance_data = [{
        'Student Name': attendance.student_name,
        'Student ID': attendance.student_number,
        'Program Code': attendance.program_code,
        'Year Level': attendance.level_code,
        'Section': attendance.section,
        'Semester': attendance.semester,
        'Course': attendance.course_name,
        'Faculty': attendance.faculty_name,
        'Time In': attendance.time_in.strftime("%Y-%m-%d %H:%M:%S") if attendance.time_in else '',
        'Time Out': attendance.time_out.strftime("%Y-%m-%d %H:%M:%S") if attendance.time_out else '',
        'Status': attendance.status
    } for attendance in attendances]

    # Create DataFrame
    df = pd.DataFrame(attendance_data)

    # Export to Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Attendance')

    output.seek(0)

    # Create response
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=attendance.xlsx"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return response


def add_header_footer_detail(canvas, doc, is_first_page, start_date=None, end_date=None):
    canvas.saveState()

    # Header Section - different logic for first page
    if is_first_page:
        # Left side logo
        logo_path = os.path.join(os.getcwd(), 'static', 'images', 'logo', 'cspc.png')
        if os.path.exists(logo_path):
            canvas.drawImage(logo_path, 65, doc.height + 40, width=50, height=50)  # Adjusted logo size and position

        # Left side text
        canvas.setFont("Helvetica", 10)
        canvas.drawString(120, doc.height + 75, "Republic of the Philippines")  # Adjusted header text position
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawString(120, doc.height + 60, "CAMARINES SUR POLYTECHNIC COLLEGES")
        canvas.setFont("Helvetica", 10)
        canvas.drawString(120, doc.height + 45, "Nabua, Camarines Sur")
        canvas.setFont("Helvetica", 6)
        canvas.drawString(70, doc.height + 30, "ISO 9001:2015")

        # Right side logo
        new_logo_path = os.path.join(os.getcwd(), 'static', 'images', 'logo', 'ccs-logo.png')
        if os.path.exists(new_logo_path):
            canvas.drawImage(new_logo_path, doc.width - 1, doc.height + 40, width=50,
                             height=50)  # Align to match left logo size

        # Right side text (College of Computer Studies)
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawString(doc.width - 215, doc.height + 60,
                          "COLLEGE OF COMPUTER STUDIES")  # Text aligned to the left of the logo

        # Horizontal line after the header
        canvas.setLineWidth(1)
        canvas.setStrokeColor(colors.black)
        canvas.line(60, doc.height + 25, doc.width + 45, doc.height + 25)  # Draw a horizontal line after the header

    else:
        # Adjusted header for the 2nd and subsequent pages
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawString(65, doc.height + 60, "CAMARINES SUR POLYTECHNIC COLLEGES, ATTENDANCE RECORD LOGS - CONTINUED")

    # Add the italic footer text
    canvas.setFont("Helvetica-Oblique", 9)  # Setting the font to italic
    canvas.drawString(60, 50, "System-generated report. No Signature required.")  # Adjust the y-position as needed

    # Footer Section with horizontal line
    canvas.setLineWidth(1)
    canvas.setStrokeColor(colors.black)
    canvas.line(60, 40, doc.width + 45, 40)  # Horizontal line for the footer
    canvas.setFont("Helvetica", 10)

    # Format the effectivity date based on the provided start and end dates
    if start_date and end_date:
        effectivity_date = f"From {start_date.strftime('%Y-%m-%d')} - To {end_date.strftime('%Y-%m-%d')}"
    elif start_date:
        effectivity_date = f"From {start_date.strftime('%Y-%m-%d')}"
    elif end_date:
        effectivity_date = f"To {end_date.strftime('%Y-%m-%d')}"
    else:
        effectivity_date = "All records across all dates."

    canvas.drawString(60, 30, effectivity_date)

    # Include current_user.faculty_details.full_name in the footer
    if current_user and hasattr(current_user, 'faculty_details') and current_user.faculty_details:
        faculty_full_name = current_user.faculty_details.full_name
        canvas.drawString(60, 20, f"Prepared by: {faculty_full_name.title()}")

    canvas.drawRightString(doc.width + 45, 30, f"Page {doc.page}")

    canvas.restoreState()


def export_pdf_detail(attendances, selected_columns, start_date=None, end_date=None):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(GOV_LEGAL),
                            leftMargin=0.75 * inch, rightMargin=0.75 * inch,
                            topMargin=1 * inch, bottomMargin=0.75 * inch)

    styles = getSampleStyleSheet()
    elements = []
    elements.append(Spacer(1, 30))  # Spacer to push the title down
    title = Paragraph("<b>ATTENDANCE RECORD LOGS</b>", styles['Title'])
    title.hAlign = 'CENTER'
    elements.append(title)
    elements.append(Spacer(1, 10))

    # Define column mappings for selected columns
    column_mapping = {
        'student_name': 'Student Name',
        'student_number': 'Student ID',
        'course': 'Subject',
        'date': 'Date',
        'program_section': 'Course & Section',
        'semester': 'Semester',
        'time_in': 'Time In',
        'time_out': 'Time Out',
        'status': 'Status'
    }

    # Create table header based on selected columns
    table_header = [column_mapping[col] for col in selected_columns]
    data = [[Paragraph(cell, styles['Normal']) for cell in table_header]]

    # Populate the table rows based on selected columns
    for attendance in attendances:
        row = []
        if 'student_name' in selected_columns:
            row.append(Paragraph(attendance.student_name.title(), styles['Normal']))
        if 'student_number' in selected_columns:
            row.append(Paragraph(attendance.student_number.title(), styles['Normal']))
        if 'course' in selected_columns:
            row.append(Paragraph(attendance.course_name.title(), styles['Normal']))
        if 'date' in selected_columns:
            row.append(
                Paragraph(attendance.date.strftime('%b. %d, %Y') if attendance.date else 'N/A', styles['Normal']))
        if 'program_section' in selected_columns:
            row.append(Paragraph(
                f'{attendance.program_code.upper()} {attendance.level_code}{attendance.section.upper() if attendance.section else ""}',
                styles['Normal']))
        if 'semester' in selected_columns:
            row.append(Paragraph(attendance.semester, styles['Normal']))
        if 'time_in' in selected_columns:
            row.append(
                Paragraph(attendance.time_in.strftime('%I:%M %p') if attendance.time_in else 'N/A', styles['Normal']))
        if 'time_out' in selected_columns:
            row.append(
                Paragraph(attendance.time_out.strftime('%I:%M %p') if attendance.time_out else 'N/A', styles['Normal']))
        if 'status' in selected_columns:
            row.append(Paragraph(attendance.status.upper(), styles['Normal']))
        data.append(row)

    # Define specific column widths for time_in, time_out, and status
    column_widths = {
        'student_name': 180,
        'student_number': 65,
        'course': 120,
        'date': 70,
        'program_section': 100,
        'semester': 90,
        'time_in': 65,  # Specific width for Time In
        'time_out': 65,  # Specific width for Time Out
        'status': 60  # Specific width for Status
    }

    # Set colWidths dynamically based on selected columns
    col_widths = [column_widths[col] for col in selected_columns]

    # Create the table with the specified column widths
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.white),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(table)

    def on_first_page(canvas, doc):
        add_header_footer_detail(canvas, doc, is_first_page=True, start_date=start_date, end_date=end_date)

    def on_later_pages(canvas, doc):
        add_header_footer_detail(canvas, doc, is_first_page=False, start_date=start_date, end_date=end_date)

    doc.build(elements, onFirstPage=on_first_page, onLaterPages=on_later_pages)

    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='detailed_attendance_records.pdf',
                     mimetype='application/pdf')


def export_excel(attendances, semester, school_year, course, year_level, program, section):
    # Resolve display names from database if IDs are provided
    semester_display = Semester.query.get(int(semester)).display_name if semester and semester.isdigit() else semester
    school_year_display = SchoolYear.query.get(
        int(school_year)).year_label if school_year and school_year.isdigit() else school_year
    course_obj = Course.query.get(int(course)) if course and course.isdigit() else None
    course_display = course_obj.course_name if course_obj else course
    course_code = course_obj.course_code if course_obj else 'N/A'
    program_display = Program.query.get(int(program)).program_code.upper() if program and program.isdigit() else program
    year_level_display = YearLevel.query.get(
        int(year_level)).level_code if year_level and year_level.isdigit() else year_level
    section_display = Section.query.get(int(section)).display_name if section and section.isdigit() else section

    # Extract unique attendance dates and sort them
    attendance_dates = sorted(set(attendance.date.date() for attendance in attendances))

    # Group attendance records by student
    student_attendance = {}
    for attendance in attendances:
        student_key = (attendance.student_number.upper(), attendance.student_name.title())
        if student_key not in student_attendance:
            student_attendance[student_key] = {}
        student_attendance[student_key][attendance.date.date()] = attendance.status.capitalize()

    # Create an Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Records"

    # Calculate the farthest column for headers to align to the right
    last_col_index = 2 + len(attendance_dates) + 1  # 2 for Student ID and Name, +1 for spacing
    last_col_letter = get_column_letter(last_col_index)

    # Add title and header information (aligned to the very right)
    ws.merge_cells(start_row=1, start_column=last_col_index, end_row=1, end_column=last_col_index)
    ws.cell(row=1, column=last_col_index).value = 'ATTENDANCE RECORDS LOGS'

    ws.merge_cells(start_row=2, start_column=last_col_index, end_row=2, end_column=last_col_index)
    ws.cell(row=2, column=last_col_index).value = f'{semester_display} {school_year_display}'

    ws.merge_cells(start_row=3, start_column=last_col_index, end_row=3, end_column=last_col_index)
    ws.cell(row=3, column=last_col_index).value = f'{course_code} - {course_display.upper()}'

    ws.merge_cells(start_row=4, start_column=last_col_index, end_row=4, end_column=last_col_index)
    ws.cell(row=4, column=last_col_index).value = f'{program_display} {year_level_display} {section_display}'

    # Add header row with student info and rotated date columns
    ws.append([])  # Empty row for spacing
    header_row = ['Student ID', 'Name'] + [date.strftime('%b/%d/%Y') for date in attendance_dates]
    ws.append(header_row)

    # Apply rotation to the attendance date headers
    for col_num in range(3, 3 + len(attendance_dates)):
        cell = ws.cell(row=6, column=col_num)
        cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')
        cell.font = Font(bold=True)

    # Map the status to a single letter (same as PDF logic)
    status_mapping = {'present': 'P', 'absent': 'A', 'late': 'L', 'excuse': 'E'}

    # Add attendance data for each student
    for (student_id, student_name), records in student_attendance.items():
        row = [student_id, student_name]  # Start with student ID and name
        for date in attendance_dates:
            status = records.get(date, 'N/A')  # Match date columns
            row.append(status_mapping.get(status.lower(), 'N/A'))
        ws.append(row)

    # Adjust column widths for better visibility
    ws.column_dimensions['A'].width = 15  # Student ID column
    ws.column_dimensions['B'].width = 25  # Name column
    for col_num in range(3, 3 + len(attendance_dates)):
        ws.column_dimensions[get_column_letter(col_num)].width = 5  # Smaller width for date columns

    # Save the workbook to a BytesIO buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # Return the file as an Excel download
    return send_file(buffer, as_attachment=True, download_name='attendance_records.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def add_header_footer(canvas, doc):
    canvas.saveState()

    # Check if it's the first page
    is_first_page = doc.page == 1

    # Header Section - different logic for first page
    if is_first_page:
        # Left side logo
        logo_path = os.path.join(os.getcwd(), 'static', 'images', 'logo', 'cspc.png')
        if os.path.exists(logo_path):
            canvas.drawImage(logo_path, 65, doc.height + 40, width=50, height=50)  # Adjusted logo size and position

        # Left side text
        canvas.setFont("Helvetica", 10)
        canvas.drawString(120, doc.height + 75, "Republic of the Philippines")  # Adjusted header text position
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawString(120, doc.height + 60, "CAMARINES SUR POLYTECHNIC COLLEGES")
        canvas.setFont("Helvetica", 10)
        canvas.drawString(120, doc.height + 45, "Nabua, Camarines Sur")
        canvas.setFont("Helvetica", 6)
        canvas.drawString(70, doc.height + 30, "ISO 9001:2015")

        # Right side logo
        new_logo_path = os.path.join(os.getcwd(), 'static', 'images', 'logo', 'ccs-logo.png')
        if os.path.exists(new_logo_path):
            canvas.drawImage(new_logo_path, doc.width - 1, doc.height + 40, width=50,
                             height=50)  # Align to match left logo size

        # Right side text (College of Computer Studies)
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawString(doc.width - 215, doc.height + 60,
                          "COLLEGE OF COMPUTER STUDIES")  # Text aligned to the left of the logo

        # Horizontal line after the header
        canvas.setLineWidth(1)
        canvas.setStrokeColor(colors.black)
        canvas.line(60, doc.height + 25, doc.width + 45, doc.height + 25)  # Draw a horizontal line after the header

    else:
        # Adjusted header for the 2nd and subsequent pages
        canvas.setFont("Helvetica-Bold", 12)
        canvas.drawString(65, doc.height + 60, "CAMARINES SUR POLYTECHNIC COLLEGES, ATTENDANCE RECORD LOGS - CONTINUED")

    # Add the italic footer text
    canvas.setFont("Helvetica-Oblique", 9)  # Setting the font to italic
    canvas.drawString(60, 50, "System-generated report. No Signature required.")  # Adjust the y-position as needed

    # Footer Section with horizontal line
    canvas.setLineWidth(1)
    canvas.setStrokeColor(colors.black)
    canvas.line(60, 40, doc.width + 45, 40)  # Horizontal line for the footer
    canvas.setFont("Helvetica", 10)

    # Include current_user.faculty_details.full_name in the footer
    if current_user and hasattr(current_user, 'faculty_details') and current_user.faculty_details:
        faculty_full_name = current_user.faculty_details.full_name
        canvas.drawString(60, 20, f"Prepared by: {faculty_full_name.title()}")

    canvas.drawRightString(doc.width + 45, 30, f"Page {doc.page}")

    canvas.restoreState()


def export_pdf(attendances, semester, school_year, course, year_level, program, section):
    # Resolve display names from database if IDs are provided
    semester_display = Semester.query.get(int(semester)).display_name if semester and semester.isdigit() else semester
    school_year_display = SchoolYear.query.get(
        int(school_year)).year_label if school_year and school_year.isdigit() else school_year
    course_obj = Course.query.get(int(course)) if course and course.isdigit() else None
    course_display = course_obj.course_name if course_obj else course
    course_code = course_obj.course_code if course_obj else 'N/A'
    program_display = Program.query.get(int(program)).program_code.upper() if program and program.isdigit() else program
    year_level_display = YearLevel.query.get(
        int(year_level)).level_code if year_level and year_level.isdigit() else year_level
    section_display = Section.query.get(int(section)).display_name if section and section.isdigit() else section

    # Extract unique attendance dates and sort them
    attendance_dates = sorted(set(attendance.date.date() for attendance in attendances))

    # Group attendance records by student (same logic as in the view)
    student_attendance = {}
    for attendance in attendances:
        student_key = (attendance.student_number.upper(), attendance.student_name.title())
        if student_key not in student_attendance:
            student_attendance[student_key] = {}
        date_key = attendance.date.strftime('%b. %d, %Y')  # Consistent date format
        student_attendance[student_key][date_key] = attendance.status.capitalize()

    # Prepare the PDF
    buffer = io.BytesIO()
    doc = BaseDocTemplate(buffer, pagesize=landscape(legal),
                          leftMargin=0.75 * inch, rightMargin=0.75 * inch,
                          topMargin=1 * inch, bottomMargin=0.75 * inch)

    # Define a frame and page template
    frame = Frame(doc.leftMargin, doc.bottomMargin, doc.width, doc.height, id='normal')
    template = PageTemplate(id='attendance_template', frames=frame, onPage=add_header_footer)
    doc.addPageTemplates([template])

    elements = []

    # Add title and header information
    elements.append(Spacer(1, 30))
    title = Paragraph("<b>ATTENDANCE RECORDS LOGS</b>", getSampleStyleSheet()['Title'])
    elements.append(title)

    # Centered course and semester info
    elements.append(Spacer(1, -10))
    course_info = Paragraph(
        f"<b>{semester_display} {school_year_display}</b><br/><b>{course_code} - {course_display.upper()}</b><br/><b>{program_display} {year_level_display}{section_display}</b>",
        getSampleStyleSheet()['Normal'])
    elements.append(course_info)

    # Helper function to format the date in a stacked format
    def stack_date(date_str):
        # Split the date into components and stack them using HTML <br/>
        return "<br/>".join(date_str)

    # Modify the table header: Student ID, Name, and Attendance Dates (Stacked format)
    stacked_dates = [stack_date(date.strftime('%b/%d/%Y')) for date in attendance_dates]

    # Convert each stacked date to a Paragraph with appropriate styling
    styles = getSampleStyleSheet()
    stacked_date_paragraphs = [Paragraph(f"<font size=12>{date}</font>", styles['Normal']) for date in stacked_dates]

    # Define a larger font style for the table header with center alignment
    header_style = ParagraphStyle(name='Header', fontSize=14, alignment=TA_CENTER, spaceBefore=50, spaceAfter=6)

    # Create the table header with bigger font for 'Student ID' and 'Name'
    table_header = [
                       Paragraph('Student ID', header_style),
                       Paragraph('Name', header_style)
                   ] + stacked_date_paragraphs

    # Initialize the table data with the header
    data = [table_header]

    # Map the status to a single letter
    status_mapping = {'present': 'P', 'absent': 'A', 'late': 'L', 'excuse': 'E'}

    # Add attendance data for each student
    for (student_id, student_name), records in student_attendance.items():
        row = [student_id, student_name]  # Start with student ID and name
        for date in attendance_dates:
            status = records.get(date.strftime('%b. %d, %Y'), 'N/A')  # Match date columns
            row.append(status_mapping.get(status.lower(), 'N/A'))
        data.append(row)

    # Set column widths: Student ID (100), Name (150), and smaller width (40) for each date column
    column_widths = [100, 150] + [20] * len(attendance_dates)

    # Create and style the table
    table = Table(data, colWidths=column_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Header background color
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),  # Header text color
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Center alignment for all cells
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Bold font for headers
        ('FONTSIZE', (0, 0), (-1, 0), 10),  # Font size for headers
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Grid for all cells
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),  # Row backgrounds
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Vertical alignment
    ]))

    # Add the table to the PDF elements
    elements.append(table)

    # Build the PDF and return it as a download
    doc.build(elements)

    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='attendance_records.pdf', mimetype='application/pdf')


@faculty_acc_bp.route('/students')
@login_required
@cspc_acc_required
@faculty_required
@check_totp_verified
def view_students():
    # Query all students joined with necessary tables and filtered by faculty_id
    students = Student.query.join(User, Student.user_id == User.id).join(
        student_course_association, Student.id == student_course_association.c.student_id
    ).join(Course, student_course_association.c.course_id == Course.id).join(
        faculty_course_association, Course.id == faculty_course_association.c.course_id
    ).join(Faculty, faculty_course_association.c.faculty_id == Faculty.id).filter(
        Faculty.id == current_user.faculty_details.id
    ).all()

    return render_template('faculty_acc/view_students.html', students=students)


@faculty_acc_bp.route('/instructor/student/<int:student_id>/schedule')
@login_required
def view_student_schedule(student_id):
    # Ensure the current user is a faculty_acc
    if current_user.role != 'faculty':
        return "You are not authorized to view this page.", 403

    student = Student.query.get_or_404(student_id)

    # Fetch the schedule details
    schedule = db.session.query(Schedule).join(Section).filter(Section.id == student.section_id).all()

    return render_template('faculty_acc/schedule.html', schedule=schedule, student=student)


@faculty_acc_bp.route('/students/detailed-attendance', methods=['GET', 'POST'])
@login_required
@cspc_acc_required
@faculty_required
@check_totp_verified
def view_detailed_attendance():
    # Check for export requests
    export_type = request.args.get('export')

    # Retrieve the selected columns for export
    selected_columns = request.args.getlist('columns')
    if not selected_columns:
        selected_columns = ['student_name', 'student_number', 'course', 'date', 'program_section',
                            'semester', 'time_in', 'time_out', 'status']  # Default columns

    faculty = Faculty.query.filter_by(user_id=current_user.id).first()
    courses = faculty.courses

    # Get the start and end date from the query parameters
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None

    # Extract values from the models to populate filter options
    semester_choices = [(sem.display_name, sem.display_name) for sem in
                        Semester.query.all()]  # Use display_name for both value and label
    school_year_choices = [(sy.id, sy.year_label) for sy in SchoolYear.query.all()]
    course_choices = [(course.id, course.course_name) for course in Course.query.all()]
    program_choices = [(prog.id, prog.program_code) for prog in Program.query.all()]
    section_choices = [(sec.id, sec.display_name) for sec in Section.query.all()]
    year_level_choices = [(yl.id, yl.level_code) for yl in YearLevel.query.all()]

    # Get filter values from the request
    semester = request.args.get('semester')
    school_year = request.args.get('school_year')
    course = request.args.get('course')
    program = request.args.get('program')
    section = request.args.get('section')

    # Build the initial query to fetch attendance records
    attendances_query = Attendance.query

    # Apply date range filtering
    if start_date:
        attendances_query = attendances_query.filter(Attendance.date >= start_date)
    if end_date:
        attendances_query = attendances_query.filter(
            Attendance.date <= end_date.replace(hour=23, minute=59, second=59))

    # Apply additional filters based on request arguments
    if semester:
        attendances_query = attendances_query.filter(Attendance.semester == semester)

    if school_year:
        attendances_query = attendances_query.join(Student).join(SchoolYear).filter(SchoolYear.id == school_year)

    if course:
        attendances_query = attendances_query.join(Course).filter(Course.id == course)

    # Fetch the results
    results = attendances_query.all()

    # Filter based on faculty name with exact or fuzzy matching
    attendances = [
        attendance for attendance in results
        if attendance.faculty_name == faculty.full_name or fuzz.ratio(attendance.faculty_name, faculty.full_name) >= 80
    ]

    print(attendances)

    # Handle export requests
    if export_type in ['excel', 'pdf']:
        if export_type == 'excel':
            return export_excel_detail(attendances)
        elif export_type == 'pdf':
            return export_pdf_detail(attendances, selected_columns, start_date=start_date, end_date=end_date)

    form = AttendanceStatusForm()

    # Handle form submission to update attendance status
    if form.validate_on_submit():
        attendance_id = request.form.get('attendance_id')
        attendance = Attendance.query.get(attendance_id)
        if attendance:
            attendance.status = form.status.data
            db.session.commit()
            flash(f'Attendance status updated successfully for {attendance.student_name.title()}', 'success')
        return redirect(url_for('faculty_acc.view_detailed_attendance'))

    # Pass the filter options and selected choices to the template
    return render_template(
        'faculty_acc/view_detailed_attendance.html',
        attendances=attendances,
        form=form,
        selected_columns=selected_columns,
        semester_choices=semester_choices,
        course_choices=course_choices,
        program_choices=program_choices,
        section_choices=section_choices,
        selected_semester=semester,
        selected_school_year=school_year,
        selected_course=course,
        selected_program=program,
        selected_section=section
    )


@faculty_acc_bp.route('/students/new_attendance', methods=['GET', 'POST'])
@login_required
@cspc_acc_required
@faculty_required
@check_totp_verified
def view_new_attendance():
    form = AttendanceStatusForm()

    # Check for export requests
    export_type = request.args.get('export')
    # Get filter values from the request
    semester = request.args.get('semester')
    school_year = request.args.get('school_year')
    course = request.args.get('course')
    program = request.args.get('program')
    year_level = request.args.get('year_level')
    section = request.args.get('section')

    # If a school year is selected, get its label
    if school_year:
        school_year_label = SchoolYear.query.filter_by(id=school_year).first().year_label
    else:
        school_year_label = None

    # If a course is selected, get its name
    if course:
        course_name = Course.query.filter_by(id=course).first().course_name
    else:
        course_name = None

    # If a program is selected, get its code
    if program:
        program_code = Program.query.filter_by(id=program).first().program_code
    else:
        program_code = None

    # If a year level is selected, get its code
    if year_level:
        year_level_code = YearLevel.query.filter_by(id=year_level).first().level_code
    else:
        year_level_code = None

    # If a section is selected, get its display name
    if section:
        section_name = Section.query.filter_by(id=section).first().display_name
    else:
        section_name = None

    faculty = Faculty.query.filter_by(user_id=current_user.id).first()

    # Check if any filters are selected
    filters_applied = any([semester, school_year, course, program, year_level, section])

    attendances = []
    attendance_dates = []
    student_attendance = {}

    if filters_applied:

        # Extract values from the models to populate filter options
        semester_choices = [(sem.display_name, sem.display_name) for sem in Semester.query.all()]
        school_year_choices = [(sy.id, sy.year_label) for sy in SchoolYear.query.all()]

        # Use distinct to get unique courses associated with the faculty
        course_choices = db.session.query(Attendance.course_id, Attendance.course_name).filter(
            Attendance.faculty_name == faculty.full_name
        ).distinct(Attendance.course_id).all()
        course_choices = [(course_id, course_name) for course_id, course_name in course_choices]

        program_choices = [(prog.id, prog.program_code) for prog in Program.query.all()]
        section_choices = [(sec.id, sec.display_name) for sec in Section.query.all()]
        year_level_choices = [(yl.id, yl.level_code) for yl in YearLevel.query.all()]

        # Create two aliases for the Student table to avoid alias conflicts
        student_alias_1 = aliased(Student)
        student_alias_2 = aliased(Student)
        student_alias_3 = aliased(Student)

        # Build the initial query to fetch attendance records
        attendances_query = Attendance.query

        # Apply additional filters based on request arguments
        if semester:
            attendances_query = attendances_query.filter(Attendance.semester == semester)

        if school_year:
            attendances_query = attendances_query.join(student_alias_1,
                                                       Attendance.student_id == student_alias_1.id).join(
                SchoolYear, student_alias_1.school_year_id == SchoolYear.id).filter(SchoolYear.id == school_year)

        if course:
            attendances_query = attendances_query.join(Course, Attendance.course_id == Course.id).filter(
                Course.id == course)

        if program:
            attendances_query = attendances_query.join(student_alias_2,
                                                       Attendance.student_id == student_alias_2.id).join(
                Program, student_alias_2.program_id == Program.id).filter(Program.id == program)

        if year_level:
            attendances_query = attendances_query.join(student_alias_3,
                                                       Attendance.student_id == student_alias_3.id).join(
                YearLevel, student_alias_3.year_level_id == YearLevel.id).filter(YearLevel.id == year_level)

        if section:
            attendances_query = attendances_query.filter(Attendance.section == section_name)

        # Fetch the results
        results = attendances_query.all()

        # Filter based on faculty name with exact or fuzzy matching
        attendances = [
            attendance for attendance in results
            if
            attendance.faculty_name == faculty.full_name or fuzz.ratio(attendance.faculty_name, faculty.full_name) >= 80
        ]

        # Extract unique attendance dates from the results
        attendance_dates = sorted(set([attendance.date.date() for attendance in attendances]))

        # Group the attendance records by student
        student_attendance = {}
        for attendance in attendances:

            student_key = (attendance.student_number, attendance.student_name)
            if student_key not in student_attendance:
                student_attendance[student_key] = {}
            student_attendance[student_key][attendance.date.date()] = attendance.status

            print(student_key)

    # Handle export requests
    if export_type in ['pdf', 'excel']:
        if export_type == 'pdf':
            return export_pdf(attendances, semester, school_year, course, program, year_level, section)
        elif export_type == 'excel':
            return export_excel(attendances, semester, school_year, course, program, year_level, section)

    # Pass the filter options and the filtered results to the template
    return render_template(
        'faculty_acc/view_new_attendance.html',
        attendances=attendances,
        attendance_dates=attendance_dates,  # Pass the attendance dates
        student_attendance=student_attendance,  # Pass the student attendance records
        form=AttendanceStatusForm(),
        semester_choices=[(sem.display_name, sem.display_name) for sem in Semester.query.all()],
        school_year_choices=[(sy.id, sy.year_label) for sy in SchoolYear.query.all()],
        course_choices=[(course_id, course_name) for course_id, course_name in db.session.query(
            Attendance.course_id, Attendance.course_name
        ).filter(Attendance.faculty_name == faculty.full_name).distinct(Attendance.course_id).all()],
        program_choices=[(prog.id, prog.program_code) for prog in Program.query.all()],
        year_level_choices=[(yl.id, yl.level_code) for yl in YearLevel.query.all()],
        section_choices=[(sec.id, sec.display_name) for sec in Section.query.all()],
        selected_semester=semester,
        selected_school_year=school_year,
        selected_course=course,
        selected_program=program,
        selected_year_level=year_level,
        selected_section=section,
        filtered_school_year=school_year_label,  # Pass the label instead of the id
        filtered_course=course_name,  # Pass the name instead of the id
        filtered_program=program_code,  # Pass the code instead of the id
        filtered_year_level=year_level_code,  # Pass the code instead of the id
        filtered_section=section_name  # Pass the display name instead of the id
    )


@faculty_acc_bp.route("/profile/<int:faculty_id>", methods=["GET", "POST"])
@login_required
@cspc_acc_required
@own_faculty_account_required
def faculty_profile(faculty_id):
    # Use the user associated with the email in the session
    faculty = Faculty.query.filter_by(id=faculty_id, user_id=current_user.id).first_or_404()
    user = faculty.user
    form = FacultyForm(faculty_id=faculty.id, obj=user)
    print(request.form)

    def clean_field(field_value):
        return " ".join(field_value.strip().split())

    def convert_to_hex(rfid_uid):
        try:
            # Try to interpret the RFID UID as an integer (which means it's in decimal)
            rfid_int = int(rfid_uid)

            # Convert the integer to a hexadecimal string
            rfid_hex = format(rfid_int, '08X')

            # Adjust the byte order (reverse for little-endian)
            rfid_hex = ''.join(reversed([rfid_hex[i:i + 2] for i in range(0, len(rfid_hex), 2)]))

            return rfid_hex
        except ValueError:
            # If it raises a ValueError, it means rfid_uid is already a valid hexadecimal string
            return rfid_uid.lower()  # Return as lowercase hex

    if request.method == 'POST':
        print(form.errors)  # This will print any validation errors in the form

    if request.method == "GET":
        # Populate the form with existing data, using `or ''` to handle None values
        form.rfid_uid.data = user.rfid_uid or ''

        form.f_name.data = user.f_name or ''
        form.l_name.data = user.l_name or ''
        form.m_name.data = user.m_name or ''

        form.gender.data = user.gender or ''

        # Populate Faculty-specific data
        form.school_id.data = faculty.faculty_number or ''
        form.department.data = faculty.faculty_department or ''
        form.designation.data = faculty.designation or ''

    if form.validate_on_submit():
        print("Form is valid")
        # Convert the RFID UID to hexadecimal before saving
        form.rfid_uid.data = convert_to_hex(form.rfid_uid.data)

        # Convert all fields to lowercase and normalize spaces
        for field in form:
            if isinstance(field.data, str):
                # Convert to lowercase
                field.data = field.data.lower()
                # Normalize spaces (reduce multiple spaces to a single space)
                field.data = re.sub(r'\s+', ' ', field.data)

        # Check for existing email,  and faculty number conflicts
        rfid_uid_exists = User.query.filter(User.rfid_uid == form.rfid_uid.data, User.id != user.id).first()
        email_exists = User.query.filter(User.email == form.email.data, User.id != user.id).first()
        faculty_number_exists = Faculty.query.filter(Faculty.faculty_number == form.school_id.data,
                                                     Faculty.user_id != user.id).first()
        if rfid_uid_exists is not None:
            flash('RFID is already in use', 'error')
        elif email_exists is not None:
            flash('Email is already in use', 'error')
        # elif username_exists is not None:
        #     flash('Username is already taken', 'error')
        elif faculty_number_exists is not None:
            flash('Faculty Number is already in use', 'error')
        else:
            from sqlalchemy.exc import IntegrityError, SQLAlchemyError

            # Inside your route function
            try:
                with db.session.begin_nested():
                    # Convert all fields to lowercase and normalize spaces
                    user.rfid_uid = clean_field(form.rfid_uid.data.lower())
                    user.f_name = clean_field(form.f_name.data.lower())
                    user.l_name = clean_field(form.l_name.data.lower())
                    user.m_name = clean_field(form.m_name.data.lower())
                    user.gender = clean_field(form.gender.data.lower())

                    # Update Faculty
                    faculty.faculty_number = clean_field(form.school_id.data.lower())
                    faculty.faculty_department = clean_field(form.department.data.lower())
                    faculty.designation = clean_field(form.designation.data.lower())

                db.session.commit()
                flash('Faculty details updated successfully!', 'success')

                # redirect to the two-factor auth page, passing username in session
                session['email'] = user.email
                return redirect(url_for('totp.two_factor_setup'))

            except IntegrityError as e:
                db.session.rollback()
                # Handle duplicate entry errors
                if "Duplicate entry" in str(e.orig):
                    field_name = str(e.orig).split("'")[3]

                    field_name_map = {
                        'rfid_uid': 'RFID',
                        'email': 'Email',
                        'faculty_number': 'Faculty Number',
                    }

                    friendly_field_name = field_name_map.get(field_name, field_name)
                    flash(f"The {friendly_field_name} you entered is already in use. Please use a different value.",
                          'error')
                else:
                    flash('An error occurred while updating the faculty details. Please try again.', 'error')

            except SQLAlchemyError as e:
                db.session.rollback()
                # Handle other SQLAlchemy errors
                flash('A database error occurred. Please try again.', 'error')

            if form.errors:
                for field, errors in form.errors.items():
                    for error in errors:
                        flash(f"Error: {error}", 'error')

    return render_template('faculty_acc/faculty_profile.html', form=form, faculty=faculty)


@faculty_acc_bp.route('/view_schedule/<int:faculty_id>')
@login_required
@cspc_acc_required
@own_faculty_account_required
@check_totp_verified
def view_schedule(faculty_id):
    faculty = Faculty.query.get_or_404(faculty_id)

    # Query schedules joined with necessary tables and filtered by faculty_id
    schedules = Schedule.query.join(Course, Schedule.course_id == Course.id) \
        .join(FacultyCourseSchedule, Schedule.id == FacultyCourseSchedule.schedule_id) \
        .join(Faculty, FacultyCourseSchedule.faculty_id == Faculty.id) \
        .join(Program, FacultyCourseSchedule.program_id == Program.id) \
        .join(YearLevel, FacultyCourseSchedule.year_level_id == YearLevel.id) \
        .join(Section, FacultyCourseSchedule.section_id == Section.id) \
        .join(Semester, FacultyCourseSchedule.semester_id == Semester.id) \
        .filter(Faculty.id == faculty_id).all()

    # Format start and end times to 12-hour format and remove SemesterEnum prefix
    for sched in schedules:
        sched.formatted_start_time = datetime.strptime(str(sched.start_time), "%H:%M:%S").strftime("%I:%M %p")
        sched.formatted_end_time = datetime.strptime(str(sched.end_time), "%H:%M:%S").strftime("%I:%M %p")
        sched.formatted_semester_name = str(sched.faculty_course_schedules[0].semester.display_name)

    return render_template('faculty_acc/view_schedule.html', faculty=faculty, schedules=schedules)
